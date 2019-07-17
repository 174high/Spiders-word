import os,sys
from openpyxl import load_workbook
from openpyxl import Workbook

def merge_file(path,output_file):


    print(path)
    print(output_file)

    fileList=os.listdir(path)
    wkbk = Workbook()

    for file in fileList:
        wb = load_workbook(path+file)
        print(wb.sheetnames[0])
 
        outsheet = wkbk.create_sheet(wb.sheetnames[0]) 
        for sheet in wb:
            print("max row=",sheet.max_row,"max column=",sheet.max_column)
            for i in  range(1,sheet.max_row+1):
                for j in range(1,sheet.max_column+1):
#                    print((outsheet.cell(row=i, column=j)))
                    outsheet.cell(row=i, column=j,value=sheet.cell(row=i, column=j).value)

    wkbk.remove(wkbk.get_sheet_by_name("Sheet"))
    wkbk.save(output_file)

def summary_log(path,output_file):

    print(path)
    print(output_file)

    fileList=os.listdir(path)
    wkbk = Workbook()
    outsheet = wkbk.create_sheet("summary") 

    outsheet.cell(row=1, column=1,value="Equipment #")
    outsheet.cell(row=1, column=2,value="Error")
    outsheet.cell(row=1, column=3,value="Emergency")
    outsheet.cell(row=1, column=4,value="Warning")
    outsheet.cell(row=1, column=5,value="Breakdown")
    outsheet.cell(row=1, column=6,value="Alarm")

    devive_num=1
    for file in fileList:
        wb = load_workbook(path+file)
        print(wb.sheetnames[0])

        Warning=0
        Alarm=0
        Breakdown=0
        Error=0
        Emergency=0

        for sheet in wb:
            print("max row=",sheet.max_row,"max column=",sheet.max_column)
            for i in  range(8,sheet.max_row+1):
                    if sheet.cell(row=i, column=1).value == "01" :
                        
                        if sheet.cell(row=i, column=2).value[0:1] != "00" :

                            print(sheet.cell(row=i, column=1).value)
                            print(sheet.cell(row=i, column=4).value)
                            if sheet.cell(row=i, column=4).value == "Warning" :
                                Warning=Warning+1    
                            elif sheet.cell(row=i, column=4).value == "Alarm" :
                                Alarm=Alarm+1
                            elif sheet.cell(row=i, column=4).value == "Breakdown" :
                                Breakdown=Breakdown+1
                            elif sheet.cell(row=i, column=4).value == "Error" :
                                Error=Error+1
                            elif sheet.cell(row=i, column=4).value == "Emergency" :
                                Emergency=Emergency+1

            devive_num=devive_num+1
            outsheet.cell(row=devive_num, column=1,value=sheet.cell(row=2, column=2).value)
            outsheet.cell(row=devive_num, column=2,value=Error)
            outsheet.cell(row=devive_num, column=3,value=Emergency)
            outsheet.cell(row=devive_num, column=4,value=Warning)
            outsheet.cell(row=devive_num, column=5,value=Breakdown)
            outsheet.cell(row=devive_num, column=6,value=Alarm)
            
    wkbk.remove(wkbk.get_sheet_by_name("Sheet"))
    wkbk.save(output_file)

def summary_event(path,output_file):

    print(path)
    print(output_file)

    fileList=os.listdir(path)
    wkbk = Workbook()
    outsheet = wkbk.create_sheet("summary") 

    outsheet.cell(row=1, column=1,value="Equipment #")
    outsheet.cell(row=1, column=2,value="Error")
    outsheet.cell(row=1, column=3,value="Symptom")
    outsheet.cell(row=1, column=4,value="Warning")
    outsheet.cell(row=1, column=5,value="Breakdown")
    outsheet.cell(row=1, column=6,value="SW change")

    devive_num=1
    for file in fileList:
        wb = load_workbook(path+file)
        print(wb.sheetnames[0])

        Warning=0
        SW_change=0
        Breakdown=0
        Error=0
        Symptom=0

        for sheet in wb:
            print("max row=",sheet.max_row,"max column=",sheet.max_column)
            for i in  range(8,sheet.max_row+1):

                print(sheet.cell(row=i, column=2).value)
                if sheet.cell(row=i, column=2).value == "Warning" :
                    Warning=Warning+1    
                elif sheet.cell(row=i, column=2).value == "SW change" :
                    SW_change=SW_change+1
                elif sheet.cell(row=i, column=2).value == "Breakdown" :
                    Breakdown=Breakdown+1
                elif sheet.cell(row=i, column=2).value == "Error" :
                    Error=Error+1
                elif sheet.cell(row=i, column=2).value == "Symptom" :
                    Symptom=Symptom+1
                    print("Symptom=",Symptom)

            devive_num=devive_num+1
            outsheet.cell(row=devive_num, column=1,value=sheet.cell(row=2, column=2).value)
            outsheet.cell(row=devive_num, column=2,value=Error)
            outsheet.cell(row=devive_num, column=3,value=Symptom)
            outsheet.cell(row=devive_num, column=4,value=Warning)
            outsheet.cell(row=devive_num, column=5,value=Breakdown)
            outsheet.cell(row=devive_num, column=6,value=SW_change)
            
    wkbk.remove(wkbk.get_sheet_by_name("Sheet"))
    wkbk.save(output_file)

if __name__ == "__main__":

     if sys.argv[1]=="log" :
         merge_file("./device-log/","./result/merge-log.xlsx")
     elif sys.argv[1]=="event" :
         merge_file("./device-event/","./result/merge-event.xlsx")
     elif sys.argv[1]=="summary-log" :
         summary_log("./device-log/","./result/summary-log.xlsx")
     elif sys.argv[1]=="summary-event" :
         summary_event("./device-event/","./result/summary-event.xlsx")









